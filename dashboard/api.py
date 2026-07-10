# dashboard/api.py
from ninja import Router
from users.auth import JWTAuth
from caveaux.models import Caveau, Bloc
from reservations.models import Reservation
from finances.models import Paiement
from concessions.models import Concession
from exhumations.models import Exhumation
from django.db.models import Sum, Count
from django.http import HttpResponse
import csv
import openpyxl
from django.utils import timezone
from datetime import timedelta
from typing import List, Dict, Any

router = Router()
auth = JWTAuth()


def _is_admin(user) -> bool:
    """Vérifie si l'utilisateur a un rôle administratif."""
    return user.role in ["ADMIN", "SECRETARIAT"]


@router.get("/stats", auth=auth)
def get_dashboard_statistics(request):
    """Récupère les statistiques du tableau de bord."""
    user = request.auth

    if not _is_admin(user):
        return {"success": False, "message": "Accès refusé"}

    total_caveaux = Caveau.objects.count()
    disponibles = Caveau.objects.filter(statut="DISPONIBLE").count()
    reserves = Caveau.objects.filter(statut="RESERVE").count()
    occupes = Caveau.objects.filter(statut="OCCUPE").count()
    inexploitables = Caveau.objects.filter(statut="INEXPLOITABLE").count()

    taux_occupation = round((occupes / total_caveaux * 100), 2) if total_caveaux > 0 else 0

    total_revenus = Paiement.objects.filter(
        statut="VALIDE"
    ).aggregate(total=Sum('montant'))['total'] or 0

    reservations_en_attente = Reservation.objects.filter(statut="EN_ATTENTE").count()
    exhumations_en_attente = Exhumation.objects.filter(statut="EN_ATTENTE").count()
    concessions_actives = Concession.objects.filter(statut="ACTIVE").count()

    return {
        "caveaux": {
            "total": total_caveaux,
            "disponibles": disponibles,
            "reserves": reserves,
            "occupes": occupes,
            "inexploitables": inexploitables,
            "taux_occupation": f"{taux_occupation}%"
        },
        "finances": {
            "total_revenus": float(total_revenus),
        },
        "reservations_en_attente": reservations_en_attente,
        "exhumations_en_attente": exhumations_en_attente,
        "concessions_actives": concessions_actives,
    }


@router.get("/occupation-par-bloc", auth=auth, response=List[Dict[str, Any]])
def get_occupation_by_bloc(request):
    """Récupère l'occupation par bloc."""
    user = request.auth

    if not _is_admin(user):
        return {"success": False, "message": "Accès refusé"}

    blocs = Bloc.objects.all()
    result = []

    for bloc in blocs:
        total = Caveau.objects.filter(bloc=bloc).count()
        occupes = Caveau.objects.filter(bloc=bloc, statut="OCCUPE").count()
        taux = round((occupes / total * 100), 2) if total > 0 else 0

        result.append({
            "bloc": bloc.nom,
            "section": bloc.section.nom,
            "total_caveaux": total,
            "occupes": occupes,
            "taux_occupation": f"{taux}%"
        })

    return result


@router.get("/revenus-par-canal", auth=auth, response=List[Dict[str, Any]])
def get_revenues_by_channel(request):
    """Récupère les revenus par canal de paiement."""
    user = request.auth

    if not _is_admin(user):
        return {"success": False, "message": "Accès refusé"}

    revenus = Paiement.objects.filter(statut="VALIDE").values('canal').annotate(
        total=Sum('montant'),
        nombre=Count('id')
    )

    return list(revenus)


@router.get("/export-csv", auth=auth)
def export_csv_report(request):
    """Exporte les données en CSV."""
    user = request.auth

    if not _is_admin(user):
        return {"success": False, "message": "Accès refusé"}

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="registre_caveaux.csv"'

    writer = csv.writer(response)
    writer.writerow(['Référence', 'Bloc', 'Section', 'Statut', 'Longueur', 'Largeur', 'Latitude', 'Longitude'])

    for caveau in Caveau.objects.select_related('bloc__section').all():
        writer.writerow([
            caveau.reference,
            caveau.bloc.nom,
            caveau.bloc.section.nom,
            caveau.statut,
            caveau.longueur,
            caveau.largeur,
            caveau.latitude,
            caveau.longitude,
        ])

    return response


@router.get("/export-excel", auth=auth)
def export_excel_report(request):
    """Exporte les données en Excel."""
    user = request.auth

    if not _is_admin(user):
        return {"success": False, "message": "Accès refusé"}

    wb = openpyxl.Workbook()

    # Feuille Caveaux
    ws1 = wb.active
    ws1.title = "Caveaux"
    ws1.append(['Référence', 'Bloc', 'Section', 'Statut', 'Longueur', 'Largeur'])
    for caveau in Caveau.objects.select_related('bloc__section').all():
        ws1.append([
            caveau.reference,
            caveau.bloc.nom,
            caveau.bloc.section.nom,
            caveau.statut,
            caveau.longueur,
            caveau.largeur,
        ])

    # Feuille Paiements
    ws2 = wb.create_sheet("Paiements")
    ws2.append(['Référence', 'Client', 'Montant', 'Canal', 'Statut', 'Date'])
    for p in Paiement.objects.select_related('client').all():
        ws2.append([
            p.reference,
            p.client.username,
            float(p.montant),
            p.canal,
            p.statut,
            p.created_at.strftime('%d/%m/%Y %H:%M'),
        ])

    # Feuille Réservations
    ws3 = wb.create_sheet("Réservations")
    ws3.append(['Client', 'Caveau', 'Défunt', 'Statut', 'Date'])
    for r in Reservation.objects.select_related('client', 'caveau').all():
        ws3.append([
            r.client.username,
            r.caveau.reference,
            r.nom_defunt,
            r.statut,
            r.created_at.strftime('%d/%m/%Y %H:%M'),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="registre_cimetiere.xlsx"'
    wb.save(response)

    return response


@router.get("/evolution-7-jours", auth=auth, response=List[Dict[str, Any]])
def get_weekly_revenue_evolution(request):
    """Récupère l'évolution des revenus sur 7 jours."""
    user = request.auth

    if not _is_admin(user):
        return {"success": False, "message": "Accès refusé"}

    today = timezone.now().date()
    resultats = []

    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        montant = Paiement.objects.filter(
            statut="VALIDE",
            created_at__date=day
        ).aggregate(total=Sum('montant'))['total'] or 0

        resultats.append({
            "jour": day.strftime("%a"),
            "date": str(day),
            "montant": float(montant)
        })

    return resultats